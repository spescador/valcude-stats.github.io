"""
Script de importación masiva de los Excel de FBM a Supabase.
Uso: python3 scripts/import_all.py
"""

import os, sys, re
try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# ── Config ───────────────────────────────────────────────────────
FOLDER  = os.path.join(os.path.dirname(os.path.dirname(__file__)))
SUPABASE_URL = 'https://hihjbwrjxenseuaqadgv.supabase.co'
ANON_KEY     = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImhpaGpid3JqeGVuc2V1YXFhZGd2Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI0NzY5MzgsImV4cCI6MjA4ODA1MjkzOH0.IIiPwFVSbx-ugEe9kbxtqRvfJ1Cm9EdRPwmL19ZFQew'

# ── Parseo ───────────────────────────────────────────────────────
def parse_minutes(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val) * 1440  # fracción de día Excel
    s = str(val).strip()
    m = re.match(r'^(\d+):(\d{2})$', s)
    if m: return int(m.group(1)) + int(m.group(2)) / 60
    return float(s) if s else 0.0

def parse_fraction(val):
    if val is None: return 0, 0
    s = str(val).strip()
    if '/' in s:
        a, _, b = s.partition('/')
        return int(a) if a.isdigit() else 0, int(b) if b.isdigit() else 0
    return 0, int(s) if s.isdigit() else 0

def parse_player_row(row):
    name = row[1]
    if not name or str(name).strip() in ('', 'Nombre', 'TOTALES'): return None
    name = str(name).strip()
    fg2m, fg2a = parse_fraction(row[4])
    fg3m, fg3a = parse_fraction(row[6])
    ftm,  fta  = parse_fraction(row[8])
    return dict(
        jersey_number = str(row[0]).strip() if row[0] else None,
        name          = name,
        minutes       = parse_minutes(row[2]),
        points        = int(row[3] or 0),
        fg2_made=fg2m, fg2_attempted=fg2a,
        fg3_made=fg3m, fg3_attempted=fg3a,
        ft_made=ftm,   ft_attempted=fta,
        reb_def       = int(row[10] or 0),
        reb_off       = int(row[11] or 0),
        reb_total     = int(row[12] or 0),
        assists       = int(row[13] or 0),
        steals        = int(row[14] or 0),
        turnovers     = int(row[15] or 0),
        blocks_made   = int(row[16] or 0),
        blocks_received=int(row[17] or 0),
        fouls_committed=int(row[18] or 0),
        fouls_received =int(row[19] or 0),
        valuation     = int(row[20] or 0),
        plus_minus    = int(row[21] or 0),
    )

def parse_excel(path, filename):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]

    match_info = {}
    header_rows, totales_rows, team_rows = [], [], []

    for i, row in enumerate(rows):
        text = ' '.join(str(c) for c in row if c)
        if 'Estadísticas' in text or 'Estadisticas' in text:
            parts = text.split(' - ')
            if len(parts) >= 5:
                # Formato: Estadísticas - Teams - Category - Org - FBM - Season
                # season está siempre en el último part
                match_info = dict(competition=parts[2].strip(), season=parts[-1].strip())
        if row[1] == 'Nombre':   header_rows.append(i)
        if row[1] == 'TOTALES':  totales_rows.append(i)
        if row[0] and not row[1] and not row[2]:
            val = str(row[0]).strip()
            if len(val) > 3 and 'FEDERACIÓN' not in val and 'Estadística' not in val and 'APP' not in val:
                team_rows.append((i, val))

    def extract_players(h_idx, t_idx):
        players = []
        for i in range(h_idx + 1, t_idx):
            if i >= len(rows): break
            p = parse_player_row(rows[i])
            if p: players.append(p)
        return players

    jornada_m = re.match(r'^(\d+)\.', filename)
    jornada = int(jornada_m.group(1)) if jornada_m else None

    t1_players = extract_players(header_rows[0], totales_rows[0]) if len(header_rows) >= 1 else []
    t2_players = extract_players(header_rows[1], totales_rows[1]) if len(header_rows) >= 2 else []

    return dict(
        source_file  = filename,
        jornada      = jornada,
        competition  = match_info.get('competition', 'CadFem1ºaño'),
        season       = match_info.get('season', '25/26'),
        team1_name   = team_rows[0][1] if len(team_rows) > 0 else 'Equipo 1',
        team2_name   = team_rows[1][1] if len(team_rows) > 1 else 'Equipo 2',
        team1_players= t1_players,
        team2_players= t2_players,
    )

# ── SQL generation ────────────────────────────────────────────────
def sq(v):
    if v is None: return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def generate_sql(parsed):
    lines = []
    c = parsed['competition']
    s = parsed['season']
    t1 = parsed['team1_name']
    t2 = parsed['team2_name']
    sf = parsed['source_file']
    j  = parsed['jornada']
    score1 = sum(p['points'] for p in parsed['team1_players'])
    score2 = sum(p['points'] for p in parsed['team2_players'])

    lines.append(f"INSERT INTO competitions(name,season,organization) VALUES({sq(c)},{sq(s)},{sq('FBM')}) ON CONFLICT(name,season,organization) DO NOTHING;")
    lines.append(f"INSERT INTO teams(name) VALUES({sq(t1)}) ON CONFLICT(name) DO NOTHING;")
    lines.append(f"INSERT INTO teams(name) VALUES({sq(t2)}) ON CONFLICT(name) DO NOTHING;")

    for team_name, players in [(t1, parsed['team1_players']), (t2, parsed['team2_players'])]:
        for p in players:
            lines.append(
                f"INSERT INTO players(name,team_id) VALUES({sq(p['name'])},(SELECT id FROM teams WHERE name={sq(team_name)})) ON CONFLICT(name,team_id) DO NOTHING;"
            )

    lines.append(f"""INSERT INTO matches(competition_id,jornada,team1_id,team2_id,team1_score,team2_score,source_file)
VALUES(
  (SELECT id FROM competitions WHERE name={sq(c)} AND season={sq(s)} AND organization={sq('FBM')}),
  {j or 'NULL'},
  (SELECT id FROM teams WHERE name={sq(t1)}),
  (SELECT id FROM teams WHERE name={sq(t2)}),
  {score1},{score2},{sq(sf)}
) ON CONFLICT(source_file) DO NOTHING;""")

    for team_name, players in [(t1, parsed['team1_players']), (t2, parsed['team2_players'])]:
        for p in players:
            cols = "match_id,player_id,team_id,jersey_number,minutes,points,fg2_made,fg2_attempted,fg3_made,fg3_attempted,ft_made,ft_attempted,reb_def,reb_off,reb_total,assists,steals,turnovers,blocks_made,blocks_received,fouls_committed,fouls_received,valuation,plus_minus"
            vals = f"""(SELECT id FROM matches WHERE source_file={sq(sf)}),
  (SELECT id FROM players WHERE name={sq(p['name'])} AND team_id=(SELECT id FROM teams WHERE name={sq(team_name)})),
  (SELECT id FROM teams WHERE name={sq(team_name)}),
  {sq(p['jersey_number'])},{p['minutes']:.4f},{p['points']},
  {p['fg2_made']},{p['fg2_attempted']},{p['fg3_made']},{p['fg3_attempted']},
  {p['ft_made']},{p['ft_attempted']},
  {p['reb_def']},{p['reb_off']},{p['reb_total']},
  {p['assists']},{p['steals']},{p['turnovers']},
  {p['blocks_made']},{p['blocks_received']},
  {p['fouls_committed']},{p['fouls_received']},
  {p['valuation']},{p['plus_minus']}"""
            lines.append(f"INSERT INTO player_stats({cols})\nVALUES({vals})\nON CONFLICT(match_id,player_id) DO NOTHING;")

    return '\n'.join(lines)

# ── Main ─────────────────────────────────────────────────────────
def main():
    files = sorted([
        f for f in os.listdir(FOLDER)
        if f.endswith('.xlsx') and not f.startswith('~') and re.match(r'^\d+\.', f)
    ])
    print(f"Encontrados {len(files)} ficheros")
    all_sql = []
    for fname in files:
        path = os.path.join(FOLDER, fname)
        print(f"  Parseando: {fname}", end=' ... ')
        try:
            parsed = parse_excel(path, fname)
            sql    = generate_sql(parsed)
            all_sql.append(sql)
            print(f"OK ({len(parsed['team1_players'])}+{len(parsed['team2_players'])} jugadoras)")
        except Exception as e:
            print(f"ERROR: {e}")

    out_path = os.path.join(FOLDER, 'scripts', 'import_data.sql')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n\n'.join(all_sql))
    print(f"\nSQL generado en: {out_path}")
    print(f"Total sentencias: {sum(s.count(';') for s in all_sql)}")

if __name__ == '__main__':
    main()
