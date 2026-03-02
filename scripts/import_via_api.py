"""
Importación masiva de Excel a Supabase via REST API.
RLS debe estar deshabilitada antes de ejecutar este script.
"""
import os, sys, re, json
from urllib.request import urlopen, Request
from urllib.parse import urljoin
from urllib.error import HTTPError

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# ── Config ───────────────────────────────────────────────────────
SUPABASE_URL = 'https://hihjbwrjxenseuaqadgv.supabase.co'
ANON_KEY     = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImhpaGpid3JqeGVuc2V1YXFhZGd2Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI0NzY5MzgsImV4cCI6MjA4ODA1MjkzOH0.IIiPwFVSbx-ugEe9kbxtqRvfJ1Cm9EdRPwmL19ZFQew'
FOLDER       = os.path.dirname(os.path.dirname(__file__))

HEADERS = {
    'apikey':        ANON_KEY,
    'Authorization': f'Bearer {ANON_KEY}',
    'Content-Type':  'application/json',
    'Prefer':        'return=representation,resolution=merge-duplicates',
}

# ── REST helpers ─────────────────────────────────────────────────
# Columnas de conflicto por tabla
CONFLICT_COLS = {
    'competitions': 'name,season,organization',
    'teams':        'name',
    'players':      'name,team_id',
    'matches':      'source_file',
    'player_stats': 'match_id,player_id',
}

def api(method, table, data):
    conflict = CONFLICT_COLS.get(table, '')
    url  = f"{SUPABASE_URL}/rest/v1/{table}"
    if conflict:
        url += f"?on_conflict={conflict}"
    body = json.dumps(data).encode()
    req  = Request(url, data=body, headers=HEADERS, method=method)
    try:
        with urlopen(req) as r:
            resp = json.loads(r.read())
            return resp if isinstance(resp, list) else [resp]
    except HTTPError as e:
        err = e.read().decode()
        raise RuntimeError(f"HTTP {e.code} on {table}: {err}")

def upsert(table, row):
    rows = api('POST', table, row)
    return rows[0]['id'] if rows else None

def upsert_many(table, rows):
    if not rows: return []
    result = api('POST', table, rows)
    return {r['name']: r['id'] for r in result} if result else {}

# ── Parseo Excel (igual que import_all.py) ───────────────────────
def parse_minutes(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val) * 1440
    s = str(val).strip()
    m = re.match(r'^(\d+):(\d{2})$', s)
    if m: return int(m.group(1)) + int(m.group(2)) / 60
    return float(s) if s else 0.0

def parse_fraction(val):
    if val is None: return 0, 0
    s = str(val).strip()
    if '/' in s:
        a, _, b = s.partition('/')
        return int(a) if a.isdigit() else 0, int(b) if b.isdigit() else 0
    return 0, int(s) if s.isdigit() else 0

def parse_player_row(row):
    name = row[1]
    if not name or str(name).strip() in ('', 'Nombre', 'TOTALES'): return None
    name = str(name).strip()
    fg2m, fg2a = parse_fraction(row[4])
    fg3m, fg3a = parse_fraction(row[6])
    ftm,  fta  = parse_fraction(row[8])
    return dict(
        jersey_number=str(row[0]).strip() if row[0] else None,
        name=name, minutes=parse_minutes(row[2]),
        points=int(row[3] or 0),
        fg2_made=fg2m, fg2_attempted=fg2a,
        fg3_made=fg3m, fg3_attempted=fg3a,
        ft_made=ftm,   ft_attempted=fta,
        reb_def=int(row[10] or 0), reb_off=int(row[11] or 0), reb_total=int(row[12] or 0),
        assists=int(row[13] or 0), steals=int(row[14] or 0), turnovers=int(row[15] or 0),
        blocks_made=int(row[16] or 0), blocks_received=int(row[17] or 0),
        fouls_committed=int(row[18] or 0), fouls_received=int(row[19] or 0),
        valuation=int(row[20] or 0), plus_minus=int(row[21] or 0),
    )

def parse_excel(path, filename):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    match_info, header_rows, totales_rows, team_rows = {}, [], [], []
    for i, row in enumerate(rows):
        text = ' '.join(str(c) for c in row if c)
        if 'Estadísticas' in text or 'Estadisticas' in text:
            parts = text.split(' - ')
            if len(parts) >= 5:
                match_info = dict(competition=parts[2].strip(), season=parts[-1].strip())
        if row[1] == 'Nombre':  header_rows.append(i)
        if row[1] == 'TOTALES': totales_rows.append(i)
        if row[0] and not row[1] and not row[2]:
            val = str(row[0]).strip()
            if len(val) > 3 and 'FEDERACIÓN' not in val and 'Estadística' not in val and 'APP' not in val:
                team_rows.append((i, val))
    def extract_players(h, t):
        return [p for p in (parse_player_row(rows[i]) for i in range(h+1, t)) if p]
    jornada_m = re.match(r'^(\d+)\.', filename)
    return dict(
        source_file=filename, jornada=int(jornada_m.group(1)) if jornada_m else None,
        competition=match_info.get('competition', 'CadFem1ºaño'),
        season=match_info.get('season', '25/26'),
        team1_name=team_rows[0][1] if len(team_rows) > 0 else 'Equipo 1',
        team2_name=team_rows[1][1] if len(team_rows) > 1 else 'Equipo 2',
        team1_players=extract_players(header_rows[0], totales_rows[0]) if len(header_rows) >= 1 else [],
        team2_players=extract_players(header_rows[1], totales_rows[1]) if len(header_rows) >= 2 else [],
    )

# ── Import ────────────────────────────────────────────────────────
def import_match(parsed):
    sf = parsed['source_file']

    # 1. Competition
    comp_id = upsert('competitions', {
        'name': parsed['competition'], 'season': parsed['season'], 'organization': 'FBM'
    })

    # 2. Teams
    t1_id = upsert('teams', {'name': parsed['team1_name']})
    t2_id = upsert('teams', {'name': parsed['team2_name']})

    # 3. Players
    player_ids = {}
    for plist, tid in [(parsed['team1_players'], t1_id), (parsed['team2_players'], t2_id)]:
        for p in plist:
            pid = upsert('players', {'name': p['name'], 'team_id': tid})
            player_ids[(p['name'], tid)] = pid

    # 4. Match
    score1 = sum(p['points'] for p in parsed['team1_players'])
    score2 = sum(p['points'] for p in parsed['team2_players'])
    match_id = upsert('matches', {
        'competition_id': comp_id, 'jornada': parsed['jornada'],
        'team1_id': t1_id, 'team2_id': t2_id,
        'team1_score': score1, 'team2_score': score2,
        'source_file': sf,
    })

    # 5. Player stats
    for plist, tid in [(parsed['team1_players'], t1_id), (parsed['team2_players'], t2_id)]:
        for p in plist:
            pid = player_ids.get((p['name'], tid))
            if not pid: continue
            upsert('player_stats', {
                'match_id': match_id, 'player_id': pid, 'team_id': tid,
                'jersey_number': p['jersey_number'],
                'minutes': round(p['minutes'], 4), 'points': p['points'],
                'fg2_made': p['fg2_made'], 'fg2_attempted': p['fg2_attempted'],
                'fg3_made': p['fg3_made'], 'fg3_attempted': p['fg3_attempted'],
                'ft_made': p['ft_made'], 'ft_attempted': p['ft_attempted'],
                'reb_def': p['reb_def'], 'reb_off': p['reb_off'], 'reb_total': p['reb_total'],
                'assists': p['assists'], 'steals': p['steals'], 'turnovers': p['turnovers'],
                'blocks_made': p['blocks_made'], 'blocks_received': p['blocks_received'],
                'fouls_committed': p['fouls_committed'], 'fouls_received': p['fouls_received'],
                'valuation': p['valuation'], 'plus_minus': p['plus_minus'],
            })

# ── Main ─────────────────────────────────────────────────────────
def main():
    files = sorted([
        f for f in os.listdir(FOLDER)
        if f.endswith('.xlsx') and not f.startswith('~') and re.match(r'^\d+\.', f)
    ])
    print(f"Importando {len(files)} partidos...")
    ok, err = 0, 0
    for fname in files:
        path = os.path.join(FOLDER, fname)
        print(f"  {fname}", end=' ... ', flush=True)
        try:
            parsed = parse_excel(path, fname)
            import_match(parsed)
            print(f"✓ ({len(parsed['team1_players'])}+{len(parsed['team2_players'])} jugadoras)")
            ok += 1
        except Exception as e:
            print(f"✗ ERROR: {e}")
            err += 1
    print(f"\nResultado: {ok} OK, {err} errores")

if __name__ == '__main__':
    main()
